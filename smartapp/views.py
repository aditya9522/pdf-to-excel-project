from django.shortcuts import render, redirect, get_object_or_404
from django.contrib import messages
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from django.urls import reverse
import textwrap
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect
from smartapp.models import FormData, STPRecord
from openpyxl import Workbook
from .utils import pdf_to_excel, STPRecord_to_excel
import os
import io
import datetime
from django.contrib.auth import authenticate, login, logout
import pandas as pd


def index(request): 
    if request.method == 'POST':
        userid = request.POST.get('userid')
        password = request.POST.get('password')
        user = authenticate(request, username = userid, password = password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Logged in Successfully.')
            return redirect('dashboard/')
        else:
            messages.error(request, 'Invalid credencials.')
            return render(request, 'index.html')
    
    return render(request, 'index.html')

def dashboard(request):
    records = FormData.objects.all().values()
    data = list(records)
    total_records = len(data)
    context = {
        'data': data[::-1],
        'total_records': total_records,
    }
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/')

    return render(request, 'dashboard.html', context)

def bulkExcel(request):
    if request.method == 'POST':
        attachmentCount = int(request.POST.get('rows'))
        excel_dir = os.path.join(settings.MEDIA_ROOT, 'Excel-Files')
        processed_dir = os.path.join(excel_dir, 'Processed')
        error_dir = os.path.join(excel_dir, 'ErrorFiles')
        combined_file_path = os.path.join(processed_dir, 'Combined-Excel-Sheet.xlsx')

        os.makedirs(processed_dir, exist_ok=True)
        os.makedirs(error_dir, exist_ok=True)

        new_data_list = []
        error_files = []

        # Load existing data if the combined file already exists
        if os.path.exists(combined_file_path):
            combined_df = pd.read_excel(combined_file_path)
            existing_stp_ids = set(combined_df['STP ID'].unique())
            existing_data = combined_df.to_dict('records')
        else:
            existing_stp_ids = set()
            existing_data = []

        # Iterate over each uploaded file
        for i in range(1, attachmentCount + 1):
            file_object = request.FILES.get(f'attachment_excel_{i}')
            is_excel = file_object.name.split('.')[-1].lower() == 'xlsx'

            if not is_excel:
                messages.warning(request, 'Please insert Excel files only to perform the action.')
                return redirect('view-stp')

            if file_object:
                file_path = os.path.join(excel_dir, file_object.name)
                
                # Save uploaded file to disk
                with open(file_path, 'wb') as destination:
                    for chunk in file_object.chunks():
                        destination.write(chunk)

                # Read Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Check for duplicate STP IDs
                if 'STP ID' in df['Field'].values:
                    stp_id = df.loc[df['Field'] == 'STP ID', 'Value'].values[0]
                    if int(stp_id) in existing_stp_ids:
                        error_files.append(file_object.name)
                        error_file_path = os.path.join(error_dir, file_object.name)
                        os.replace(file_path, error_file_path)
                        continue
                    existing_stp_ids.add(stp_id)

                # Transform DataFrame to a dictionary and add the 'Sheet Name' column
                row_dict = {'Sheet Name': file_object.name}
                for _, row in df.iterrows():
                    row_dict[row['Field']] = row['Value']

                new_data_list.append(row_dict)

                # Move processed file to the processed directory
                processed_file_path = os.path.join(processed_dir, file_object.name)
                os.replace(file_path, processed_file_path)

        # Combine new data with existing data
        combined_data = existing_data + new_data_list

        # Convert the list of dictionaries to a DataFrame
        if combined_data:
            combined_df = pd.DataFrame(combined_data)

            # Write combined DataFrame to the existing Excel file
            combined_df.to_excel(combined_file_path, index=False)

            if os.path.exists(combined_file_path):
                with open(combined_file_path, 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=Combined-Excel-Sheet.xlsx'
                return response
        
        messages.success(request, 'Files processed successfully.')

    return redirect('view-stp')

def bulkPDF(request):
    if request.method == 'POST':
        attachmentCount = int(request.POST.get('rows'))
        pdf_dir = os.path.join(settings.MEDIA_ROOT, 'STP-Files')
        processed_dir = os.path.join(pdf_dir, 'Processed')
        error_dir = os.path.join(pdf_dir, 'Error')
        
        os.makedirs(processed_dir, exist_ok=True)
        os.makedirs(error_dir, exist_ok=True)
        
        for i in range(1, attachmentCount + 1):
            file_object = request.FILES.get(f'attachment_pdf_{i}')
            is_pdf = file_object.name.split('.')[-1].lower() == 'pdf'

            if not is_pdf:
                messages.warning(request, 'Please insert PDF files only to perform the action.')
                return redirect('view-stp')
            
            if file_object:
                file_path = os.path.join(pdf_dir, file_object.name)
                counter = 1

                # Ensure unique file path
                while os.path.exists(file_path):
                    base, ext = os.path.splitext(file_path)
                    file_path = f"{base} ({counter}){ext}"
                    counter += 1

                # Save the file temporarily
                with open(file_path, 'wb') as temp_pdf:
                    for chunk in file_object.chunks():
                        temp_pdf.write(chunk)
            
                workbook, stpid = pdf_to_excel(file_path)
                # Remove the temporary PDF file
                os.remove(file_path)

                excel_path = file_path.replace('.pdf', '.xlsx')
                counter = 1

                # Ensure unique excel path for processed/error directories
                while os.path.exists(excel_path):
                    base, ext = os.path.splitext(excel_path)
                    excel_path = f"{base} ({counter}){ext}"
                    counter += 1

                workbook.save(excel_path)

                # Check if the STP ID exists in the STPRecord model
                if not STPRecord.objects.filter(stp_id=int(stpid)).exists():
                    # If STP ID does not exist, store data in STPRecord model
                    df = pd.read_excel(excel_path, header=None, index_col=0)
                    data = df[1].to_dict()

                    def parse_date(date):
                        if pd.isna(date):
                            return None
                        return str(pd.to_datetime(date).date())
                    def clean_data(value):
                        return '' if pd.isna(value) else value

                    stp_record = STPRecord(
                        file_name = file_object.name,
                        stp_id = int(stpid),
                        product_name = clean_data(data.get('Product Name', '')),
                        batch_number = clean_data(data.get('Batch Number', '')),
                        manufacture_date = parse_date(data.get('Manufacture Date', None)),
                        expiry_date = parse_date(data.get('Expiry Date', None)),
                        active_ingredient_concentration = clean_data(data.get('Active Ingredient Concentration', '')),
                        capsule_size = clean_data(data.get('Capsule Size', '')),
                        dissolution_test = clean_data(data.get('Dissolution Test', '')),
                        hardness_test = clean_data(data.get('Hardness Test', '')),
                        moisture_content = clean_data(data.get('Moisture Content', '')),
                        dosage_unit_uniformity = clean_data(data.get('Uniformity of Dosage Unit', '')),
                        appearance = clean_data(data.get('Appearance', '')),
                        packaging_integrity = clean_data(data.get('Packaging Integrity', '')),
                        storage_conditions = clean_data(data.get('Storage Conditions', '')),
                        stability_testing = clean_data(data.get('Stability Testing', ''))
                    )
                    stp_record.save()

                    processed_path = os.path.join(processed_dir, os.path.basename(excel_path))
                    counter = 1
                    while os.path.exists(processed_path):
                        base, ext = os.path.splitext(processed_path)
                        processed_path = f"{base} ({counter}){ext}"
                        counter += 1
                    os.rename(excel_path, processed_path)
                else:
                    # If STP ID exists, save Excel sheet in error directory
                    error_path = os.path.join(error_dir, os.path.basename(excel_path))
                    counter = 1
                    while os.path.exists(error_path):
                        base, ext = os.path.splitext(error_path)
                        error_path = f"{base} ({counter}){ext}"
                        counter += 1
                    os.rename(excel_path, error_path)

        # Now storing the STPRecords to the excel sheet
        try:
            STPRecord_to_excel()
        except:
            messages.error(request, 'The process cannot access the file because it is being used by another process.')
            return redirect('view-stp')

        messages.success(request, 'Files has been processed successfully.')
        return redirect('view-stp')

    return redirect('view-stp')

def update(request, stp_id):
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/')
    
    if request.method == 'POST':
        form = get_object_or_404(FormData, stp_id=stp_id)
        form.product_name = request.POST.get('product_name', '')
        form.batch_number = request.POST.get('batch_number', '')
        form.manufacture_date = request.POST.get('manufacture_date', None) or None
        form.expiry_date = request.POST.get('expiry_date', None) or None
        form.active_ingredient_concentration = request.POST.get('active_ingredient_concentration', '')
        form.capsule_size = request.POST.get('capsule_size', '')
        form.dissolution_test = request.POST.get('dissolution_test', '')
        form.hardness_test = request.POST.get('hardness_test', '')
        form.moisture_content = request.POST.get('moisture_content', '')
        form.dosage_unit_uniformity = request.POST.get('dosage_unit_uniformity', '')
        form.appearance = request.POST.get('appearance', '')
        form.packaging_integrity = request.POST.get('packaging_integrity', '')
        form.storage_conditions = request.POST.get('storage_conditions', '')
        form.stability_testing = request.POST.get('stability_testing', '')

        if not form.product_name:
            messages.warning(request, "The product name is required!")
            update_url = reverse('update', args=[stp_id])
            return redirect(update_url)
        
        form.save()
        messages.success(request, "Record updated successfully.")
        return redirect('dashboard')
    
    try:
        form = FormData.objects.get(stp_id=stp_id) 
    except:
        # If stp_id is incorrect
        messages.error(request, 'Record not found. Please check the STP ID.')
        return redirect('dashboard')
    
    context = {
        'form': form
    }

    return render(request, 'edit.html', context)

def logoutUser(request):
    logout(request)
    return redirect('/')

def form(request):
    if FormData.objects.last():
        stp_id = FormData.objects.last().stp_id + 1
    else:
        stp_id =  1

    context = {
        'stp_id': stp_id
    }
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/', context)

    if request.method == 'POST':
        stp_id = request.POST.get('stp_id')
        product_name = request.POST.get('product_name')
        batch_number = request.POST.get('batch_number')
        manufacture_date = request.POST.get('manufacture_date', None) or None
        expiry_date = request.POST.get('expiry_date', None) or None
        active_ingredient_concentration = request.POST.get('active_ingredient_concentration')
        capsule_size = request.POST.get('capsule_size')
        dissolution_test = request.POST.get('dissolution_test')
        hardness_test = request.POST.get('hardness_test')
        moisture_content = request.POST.get('moisture_content')
        dosage_unit_uniformity = request.POST.get('dosage_unit_uniformity')
        appearance = request.POST.get('appearance')
        packaging_integrity = request.POST.get('storage_conditions')
        storage_conditions = request.POST.get('moisture_content')
        stability_testing = request.POST.get('stability_testing')     

        if not product_name:
            messages.warning(request, "The product name is required!")
            return redirect('/form/')

        form = FormData(
            stp_id=stp_id, 
            product_name=product_name,
            batch_number=batch_number,
            manufacture_date=manufacture_date,
            expiry_date=expiry_date,
            active_ingredient_concentration=active_ingredient_concentration,
            capsule_size=capsule_size,
            dissolution_test=dissolution_test,
            hardness_test=hardness_test,
            moisture_content=moisture_content,
            dosage_unit_uniformity=dosage_unit_uniformity,
            appearance=appearance,
            packaging_integrity=packaging_integrity,
            storage_conditions=storage_conditions,
            stability_testing=stability_testing,
        )
        form.save()
        messages.success(request, "Record created successfully.")
        return redirect('/dashboard/')

    return render(request, 'form.html', context)

def stpForm(request):
    if STPRecord.objects.last():
        stp_id = STPRecord.objects.last().stp_id + 1
    else:
        stp_id =  1

    context = {
        'stp_id': stp_id
    }
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/', context)

    if request.method == 'POST':
        file_name = request.POST.get('file_name')
        stp_id = request.POST.get('stp_id')
        product_name = request.POST.get('product_name')
        batch_number = request.POST.get('batch_number')
        manufacture_date = request.POST.get('manufacture_date', None) or None
        expiry_date = request.POST.get('expiry_date', None) or None
        active_ingredient_concentration = request.POST.get('active_ingredient_concentration')
        capsule_size = request.POST.get('capsule_size')
        dissolution_test = request.POST.get('dissolution_test')
        hardness_test = request.POST.get('hardness_test')
        moisture_content = request.POST.get('moisture_content')
        dosage_unit_uniformity = request.POST.get('dosage_unit_uniformity')
        appearance = request.POST.get('appearance')
        packaging_integrity = request.POST.get('storage_conditions')
        storage_conditions = request.POST.get('moisture_content')
        stability_testing = request.POST.get('stability_testing')     

        if not product_name:
            messages.warning(request, "The product name is required!")
            return redirect('/stp-form/')

        stpform = STPRecord(
            file_name=file_name,
            stp_id=stp_id, 
            product_name=product_name,
            batch_number=batch_number,
            manufacture_date=manufacture_date,
            expiry_date=expiry_date,
            active_ingredient_concentration=active_ingredient_concentration,
            capsule_size=capsule_size,
            dissolution_test=dissolution_test,
            hardness_test=hardness_test,
            moisture_content=moisture_content,
            dosage_unit_uniformity=dosage_unit_uniformity,
            appearance=appearance,
            packaging_integrity=packaging_integrity,
            storage_conditions=storage_conditions,
            stability_testing=stability_testing,
        )
        stpform.save()

        # creating excel sheet
        STPRecord_to_excel()
        messages.success(request, "STP added successfully.")
        return redirect('/view-stp/')

    return render(request, 'stp-form.html', context)

def viewSTP(request):
    try:
        records = STPRecord.objects.all().values()
        data_list = list(records)
    except Exception as e:
        data_list = []

    media_url = request.build_absolute_uri(settings.MEDIA_URL)
    context = {
        'data': data_list,
        'media_url': media_url
    }
    return render(request, 'view-stp.html', context)

def display(request, stp_id):
    if not request.user.is_authenticated:
        messages.warning(request, 'Login required!')
        return redirect('/')

    try:
        form = FormData.objects.get(stp_id=stp_id) 
    except:
        # If stp_id is incorrect
        messages.error(request, 'Record not found. Please check the STP ID.')
        return redirect('dashboard')
    
    context = {
        'form': form
    }

    return render(request, 'view.html', context)

def generate_pdf(request, stp_id):
    form = FormData.objects.get(stp_id=stp_id)
    data = {
        'STP ID': form.stp_id,
        'Product Name': form.product_name,
        'Batch Number': form.batch_number,
        'Manufacture Date': form.manufacture_date,
        'Expiry Date': form.expiry_date,
        'Active Ingredient Concentration': form.active_ingredient_concentration,
        'Capsule Size': form.capsule_size,
        'Dissolution Test': form.dissolution_test,
        'Hardness Test': form.hardness_test,
        'Moisture Content': form.moisture_content,
        'Uniformity of Dosage Unit': form.dosage_unit_uniformity,
        'Appearance': form.appearance,
        'Packaging Integrity': form.packaging_integrity,
        'Storage Conditions': form.storage_conditions,
        'Stability Testing': form.stability_testing,
    }

    for key, value in data.items():
        if isinstance(value, datetime.date):
            data[key] = value.strftime('%Y-%m-%d')

    for key, value in data.items():
        if value is None:
            data[key] = ""

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="form_{stp_id}.pdf"'

    # Create PDF document
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Define margins
    margin = 60
    usable_width = width - 2 * margin

    # Set the title
    p.setFont("Helvetica-Bold", 16)
    title = "Standard Test Procedures (STPs) for Pharmaceutical Tablets"
    title_width = p.stringWidth(title, "Helvetica-Bold", 16)
    p.drawString((width - title_width) / 2, height - margin, title)  # Centered at the top

    # Set initial y coordinate for content
    y = height - 2 * margin

    # Loop through data and add to PDF
    for key, value in data.items():
        wrapped_value = textwrap.wrap(str(value), width=int(usable_width / 5))  # Adjust width as needed
        needed_space = 20 + (15 * len(wrapped_value)) + 10  # Heading height + content lines height + additional space

        # Check if there's enough space for both the heading and its content
        if y < needed_space:  # If not enough space, start a new page
            p.showPage()  # Create a new page
            p.setFont("Helvetica-Bold", 16)
            p.drawString((width - title_width) / 2, height - margin, title)  # Centered at the top of the new page
            y = height - 2 * margin  # Reset y coordinate for new page

        # Draw heading
        p.setFont("Helvetica-Bold", 14)
        p.drawString(margin, y, f"{key}:")
        y -= 20

        # Draw content
        p.setFont("Helvetica", 12)
        for line in wrapped_value:
            if y < 20:  # Check if we need to add a new page
                p.showPage()  # Create a new page
                p.setFont("Helvetica-Bold", 16)
                p.drawString((width - title_width) / 2, height - margin, title)  # Centered at the top of the new page
                y = height - 2 * margin  # Reset y coordinate for new page
                p.setFont("Helvetica-Bold", 14)
                p.drawString(margin, y, f"{key}:")
                y -= 20

            p.drawString(margin, y, line)
            y -= 15  # Move y coordinate up for next row of content

        y -= 10  # Additional space between sections

    p.drawString(width / 2, margin / 2, "")
    p.save()
    return response

def generate_excel(request, stp_id):
    form = FormData.objects.get(stp_id=stp_id)
    data = {
        'STP ID': form.stp_id,
        'Product Name': form.product_name,
        'Batch Number': form.batch_number,
        'Manufacture Date': form.manufacture_date,
        'Expiry Date': form.expiry_date,
        'Active Ingredient Concentration': form.active_ingredient_concentration,
        'Capsule Size': form.capsule_size,
        'Dissolution Test': form.dissolution_test,
        'Hardness Test': form.hardness_test,
        'Moisture Content': form.moisture_content,
        'Uniformity of Dosage Unit': form.dosage_unit_uniformity,
        'Appearance': form.appearance,
        'Packaging Integrity': form.packaging_integrity,
        'Storage Conditions': form.storage_conditions,
        'Stability Testing': form.stability_testing,
    }

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "STP Data"

    # Set headers
    worksheet.cell(row=1, column=1, value="Field")
    worksheet.cell(row=1, column=2, value="Value")

    # Write data to Excel
    for index, (field, value) in enumerate(data.items(), start=2):
        worksheet.cell(row=index, column=1, value=field)  # Field
        worksheet.cell(row=index, column=2, value=value)  # Value

    # Adjust column width
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 60

    # Save the workbook to a BytesIO object
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Create HTTP response with Excel file
    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="form_data.xlsx"'
    
    return response

def export_excel(request, stp_id):
    if request.method == 'POST' and request.FILES.get('pdf'):
        pdf_file = request.FILES['pdf']
        is_pdf = pdf_file.name.split('.')[-1].lower() == 'pdf'

        if not is_pdf:
            messages.warning(request, 'Please insert PDF file only!')
            return HttpResponseRedirect(reverse('display', args=[stp_id]))
        
        # Save the uploaded PDF file temporarily
        temp_dir = os.path.join(settings.BASE_DIR, 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_pdf_path = os.path.join(temp_dir, pdf_file.name)
        
        with open(temp_pdf_path, 'wb') as temp_pdf:
            for chunk in pdf_file.chunks():
                temp_pdf.write(chunk)
        
        # Convert PDF to Excel
        workbook, _ = pdf_to_excel(temp_pdf_path)
        
        # Remove the temporary PDF file
        os.remove(temp_pdf_path)
        
        # Create HTTP response with Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{pdf_file.name.split(".")[0]}.xlsx"'
        
        # Save the workbook to response
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response.write(excel_file.getvalue())
        
        return response
    
    messages.warning(request, 'Please attach the file to perform the action.')
    return  HttpResponseRedirect(reverse('display', args=[stp_id]))

def custom_404(request, exception):
    return render(request, '404.html', status=404)

def custom_500(request):
    return render(request, '500.html', status=500)

