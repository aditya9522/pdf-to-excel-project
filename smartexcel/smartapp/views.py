from django.shortcuts import render, redirect, get_object_or_404
from django.http import Http404
from django.contrib import messages
from reportlab.lib.pagesizes import A4
from reportlab.pdfgen import canvas
from reportlab.lib.units import inch
from django.urls import reverse
import textwrap
from django.conf import settings
from django.http import HttpResponse, HttpResponseRedirect
from smartapp.models import FormData
from openpyxl import Workbook
from .utils import pdf_to_excel
import os
import io
import datetime
from django.contrib.auth import authenticate, login, logout
from django.utils import timezone
import pandas as pd


def index(request): 
    if request.method == 'POST':
        userid = request.POST.get('userid')
        password = request.POST.get('password')
        user = authenticate(request, username = userid, password = password)
        if user is not None:
            login(request, user)
            messages.success(request, 'Logged in Successfully.')
            return redirect('dashboard/')
        else:
            messages.error(request, 'Invalid credencials.')
            return render(request, 'index.html')
    
    return render(request, 'index.html')

def dashboard(request):
    records = FormData.objects.all().values()
    data = list(records)
    total_records = len(data)
    context = {
        'data': data[::-1],
        'total_records': total_records,
    }
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/')

    return render(request, 'dashboard.html', context)

def bulkExcel(request):
    if request.method == 'POST':
        attachmentCount = int(request.POST.get('rows'))
        excel_dir = os.path.join(settings.BASE_DIR, 'Excel-Files')
        processed_dir = os.path.join(excel_dir, 'processed')
        error_dir = os.path.join(excel_dir, 'ErrorFiles')
        os.makedirs(processed_dir, exist_ok=True)
        os.makedirs(error_dir, exist_ok=True)

        # List to store dictionaries representing each row of the combined DataFrame
        data_list = []
        stp_ids = set()
        error_files = []

        # Iterate over each uploaded file
        for i in range(1, attachmentCount + 1):
            file_object = request.FILES.get(f'attachment_excel_{i}')
            if file_object:
                file_path = os.path.join(excel_dir, file_object.name)
                
                # Save uploaded file to disk
                with open(file_path, 'wb') as destination:
                    for chunk in file_object.chunks():
                        destination.write(chunk)

                # Read Excel file into a DataFrame
                df = pd.read_excel(file_path)

                # Check for duplicate STP IDs
                if 'STP ID' in df['Field'].values:
                    stp_id = df.loc[df['Field'] == 'STP ID', 'Value'].values[0]
                    if stp_id in stp_ids:
                        error_files.append(file_object.name)
                        error_file_path = os.path.join(error_dir, file_object.name)
                        os.replace(file_path, error_file_path)
                        continue
                    stp_ids.add(stp_id)

                # Transform DataFrame to a dictionary and add the 'Sheet Name' column
                row_dict = {'Sheet Name': file_object.name}
                for _, row in df.iterrows():
                    row_dict[row['Field']] = row['Value']
                
                data_list.append(row_dict)

                # Move processed file to the processed directory
                processed_file_path = os.path.join(processed_dir, file_object.name)
                os.replace(file_path, processed_file_path)
        
        # Convert the list of dictionaries to a DataFrame
        if data_list:
            combined_df = pd.DataFrame(data_list)

            # Write combined DataFrame to a new Excel file
            combined_file_path = os.path.join(processed_dir, 'Combined-Excel-Sheet.xlsx')
            combined_df.to_excel(combined_file_path, index=False)

            if os.path.exists(combined_file_path):
                with open(combined_file_path, 'rb') as f:
                    response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                    response['Content-Disposition'] = 'attachment; filename=Combined-Excel-Sheet.xlsx'
                return response

    return redirect('dashboard')

def bulkPDF(request):
    if request.method == 'POST':
        attachmentCount = int(request.POST.get('rows'))
        pdf_dir = os.path.join(settings.BASE_DIR, 'PDF-Files')
        processed_dir = os.path.join(pdf_dir, 'Processed')
        error_dir = os.path.join(pdf_dir, 'Error')
        os.makedirs(processed_dir, exist_ok=True)
        os.makedirs(error_dir, exist_ok=True)
        data_list = []
        stp_ids = set()
        duplicate_files = []

        for i in range(1, attachmentCount + 1):
            file_object = request.FILES.get(f'attachment_pdf_{i}')
            if file_object:
                file_path = os.path.join(pdf_dir, file_object.name)

                # Saving file temporarily
                with open(file_path, 'wb') as temp_pdf:
                    for chunk in file_object.chunks():
                        temp_pdf.write(chunk)
            
                workbook, stpid = pdf_to_excel(file_path)
                # Remove the temporary PDF file
                os.remove(file_path)

                excel_path = file_path.replace('.pdf', '.xlsx')
                workbook.save(excel_path)

                if stpid in stp_ids:
                    duplicate_files.append(excel_path)
                    os.makedirs(error_dir, exist_ok=True)
                    os.rename(excel_path, os.path.join(error_dir, os.path.basename(excel_path)))
                else:
                    stp_ids.add(stpid)
                    df = pd.read_excel(excel_path)
                    row_dict = {'File Name': file_object.name}
                    for _, row in df.iterrows():
                        row_dict[row['Field']] = row['Value']
                    
                    data_list.append(row_dict)
                    os.rename(excel_path, os.path.join(processed_dir, os.path.basename(excel_path)))

        # Convert the list of dictionaries to a DataFrame
        combined_df = pd.DataFrame(data_list)

        # Write combined DataFrame to a new Excel file
        combined_file_path = os.path.join(processed_dir, 'Combined-Data-Sheet.xlsx')
        combined_df.to_excel(combined_file_path, index=False)

        if os.path.exists(combined_file_path):
            with open(combined_file_path, 'rb') as f:
                response = HttpResponse(f.read(), content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
                response['Content-Disposition'] = 'attachment; filename=Combined-Data-Sheet.xlsx'
            return response

    return redirect('dashboard')

def update(request, stp_id):
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/')
    
    if request.method == 'POST':
        form = get_object_or_404(FormData, stp_id=stp_id)
        form.product_name = request.POST.get('product_name', '')
        form.batch_number = request.POST.get('batch_number', '')
        form.manufacture_date = request.POST.get('manufacture_date', None)
        form.expiry_date = request.POST.get('expiry_date', None)
        form.active_ingredient_concentration = request.POST.get('active_ingredient_concentration', '')
        form.capsule_size = request.POST.get('capsule_size', '')
        form.dessolution_test = request.POST.get('dessolution_test', '')
        form.hardness_test = request.POST.get('hardness_test', '')
        form.moisture_content = request.POST.get('moisture_content', '')
        form.dosage_unit_uniformiry = request.POST.get('dosage_unit_uniformiry', '')
        form.appearance = request.POST.get('appearance', '')
        form.packaging_integrity = request.POST.get('packaging_integrity', '')
        form.storage_conditions = request.POST.get('storage_conditions', '')
        form.stability_testing = request.POST.get('stability_testing', '')

        if not form.product_name:
            messages.warning(request, "The product name is required!")
            update_url = reverse('update', args=[stp_id])
            return redirect(update_url)
        
        form.save()
        messages.success(request, "Record updated successfully.")
        return redirect('dashboard')
    
    try:
        form = FormData.objects.get(stp_id=stp_id) 
    except:
        # If stp_id is incorrect
        messages.error(request, 'Record not found. Please check the STP ID.')
        return redirect('dashboard')
    
    context = {
        'form': form
    }

    return render(request, 'edit.html', context)

def logoutUser(request):
    logout(request)
    return redirect('/')

def form(request):
    if FormData.objects.last():
        stp_id = FormData.objects.last().stp_id + 1
    else:
        stp_id =  1

    context = {
        'stp_id': stp_id
    }
    if not request.user.is_authenticated:
        messages.warning(request, "Login Required!")
        return redirect('/', context)

    if request.method == 'POST':
        stp_id = request.POST.get('stp_id')
        product_name = request.POST.get('product_name')
        batch_number = request.POST.get('batch_number')
        manufacture_date = request.POST.get('manufacture_date', None) or None
        expiry_date = request.POST.get('expiry_date', None) or None
        active_ingredient_concentration = request.POST.get('active_ingredient_concentration')
        capsule_size = request.POST.get('capsule_size')
        dessolution_test = request.POST.get('dessolution_test')
        hardness_test = request.POST.get('hardness_test')
        moisture_content = request.POST.get('moisture_content')
        dosage_unit_uniformiry = request.POST.get('dosage_unit_uniformiry')
        appearance = request.POST.get('appearance')
        packaging_integrity = request.POST.get('storage_conditions')
        storage_conditions = request.POST.get('moisture_content')
        stability_testing = request.POST.get('stability_testing')     

        if not product_name:
            messages.warning(request, "The product name is required!")
            return redirect('/form/')

        form = FormData(
            stp_id=stp_id, 
            product_name=product_name,
            batch_number=batch_number,
            manufacture_date=manufacture_date,
            expiry_date=expiry_date,
            active_ingredient_concentration=active_ingredient_concentration,
            capsule_size=capsule_size,
            dessolution_test=dessolution_test,
            hardness_test=hardness_test,
            moisture_content=moisture_content,
            dosage_unit_uniformiry=dosage_unit_uniformiry,
            appearance=appearance,
            packaging_integrity=packaging_integrity,
            storage_conditions=storage_conditions,
            stability_testing=stability_testing,
        )
        form.save()
        messages.success(request, "Record created successfully.")
        return redirect('/dashboard/')

    return render(request, 'form.html', context)

def display(request, stp_id):
    if not request.user.is_authenticated:
        messages.warning(request, 'Login required!')
        return redirect('/')

    try:
        form = FormData.objects.get(stp_id=stp_id) 
    except:
        # If stp_id is incorrect
        messages.error(request, 'Record not found. Please check the STP ID.')
        return redirect('dashboard')
    
    context = {
        'form': form
    }

    return render(request, 'view.html', context)

def generate_pdf(request, stp_id):
    form = FormData.objects.get(stp_id=stp_id)
    data = {
        'STP ID': form.stp_id,
        'Product Name': form.product_name,
        'Batch Number': form.batch_number,
        'Manufacture Date': form.manufacture_date,
        'Expiry Date': form.expiry_date,
        'Active Ingredient Concentration': form.active_ingredient_concentration,
        'Capsule Size': form.capsule_size,
        'Dissolution Test': form.dessolution_test,
        'Hardness Test': form.hardness_test,
        'Moisture Content': form.moisture_content,
        'Uniformity of Dosage Unit': form.dosage_unit_uniformiry,
        'Appearance': form.appearance,
        'Packaging Integrity': form.packaging_integrity,
        'Storage Conditions': form.storage_conditions,
        'Stability Testing': form.stability_testing,
    }

    for key, value in data.items():
        if isinstance(value, datetime.date):
            data[key] = value.strftime('%Y-%m-%d')

    for key, value in data.items():
        if value is None:
            data[key] = ""

    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="form_{stp_id}.pdf"'

    # Create PDF document
    p = canvas.Canvas(response, pagesize=A4)
    width, height = A4

    # Define margins
    margin = 60
    usable_width = width - 2 * margin

    # Set the title
    p.setFont("Helvetica-Bold", 16)
    title = "Standard Test Procedures (STPs) for Pharmaceutical Tablets"
    title_width = p.stringWidth(title, "Helvetica-Bold", 16)
    p.drawString((width - title_width) / 2, height - margin, title)  # Centered at the top

    # Set initial y coordinate for content
    y = height - 2 * margin

    # Loop through data and add to PDF
    for key, value in data.items():
        wrapped_value = textwrap.wrap(str(value), width=int(usable_width / 5))  # Adjust width as needed
        needed_space = 20 + (15 * len(wrapped_value)) + 10  # Heading height + content lines height + additional space

        # Check if there's enough space for both the heading and its content
        if y < needed_space:  # If not enough space, start a new page
            p.showPage()  # Create a new page
            p.setFont("Helvetica-Bold", 16)
            p.drawString((width - title_width) / 2, height - margin, title)  # Centered at the top of the new page
            y = height - 2 * margin  # Reset y coordinate for new page

        # Draw heading
        p.setFont("Helvetica-Bold", 14)
        p.drawString(margin, y, f"{key}:")
        y -= 20

        # Draw content
        p.setFont("Helvetica", 12)
        for line in wrapped_value:
            if y < 20:  # Check if we need to add a new page
                p.showPage()  # Create a new page
                p.setFont("Helvetica-Bold", 16)
                p.drawString((width - title_width) / 2, height - margin, title)  # Centered at the top of the new page
                y = height - 2 * margin  # Reset y coordinate for new page
                p.setFont("Helvetica-Bold", 14)
                p.drawString(margin, y, f"{key}:")
                y -= 20

            p.drawString(margin, y, line)
            y -= 15  # Move y coordinate up for next row of content

        y -= 10  # Additional space between sections

    p.drawString(width / 2, margin / 2, "")
    p.save()
    return response

def generate_excel(request, stp_id):
    form = FormData.objects.get(stp_id=stp_id)
    data = {
        'STP ID': form.stp_id,
        'Product Name': form.product_name,
        'Batch Number': form.batch_number,
        'Manufacture Date': form.manufacture_date,
        'Expiry Date': form.expiry_date,
        'Active Ingredient Concentration': form.active_ingredient_concentration,
        'Capsule Size': form.capsule_size,
        'Dissolution Test': form.dessolution_test,
        'Hardness Test': form.hardness_test,
        'Moisture Content': form.moisture_content,
        'Uniformity of Dosage Unit': form.dosage_unit_uniformiry,
        'Appearance': form.appearance,
        'Packaging Integrity': form.packaging_integrity,
        'Storage Conditions': form.storage_conditions,
        'Stability Testing': form.stability_testing,
    }

    # Create a new Excel workbook and add a worksheet
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "STP Data"

    # Set headers
    worksheet.cell(row=1, column=1, value="Field")
    worksheet.cell(row=1, column=2, value="Value")

    # Write data to Excel
    for index, (field, value) in enumerate(data.items(), start=2):
        worksheet.cell(row=index, column=1, value=field)  # Field
        worksheet.cell(row=index, column=2, value=value)  # Value

    # Adjust column width
    worksheet.column_dimensions['A'].width = 30
    worksheet.column_dimensions['B'].width = 60

    # Save the workbook to a BytesIO object
    excel_file = io.BytesIO()
    workbook.save(excel_file)
    excel_file.seek(0)

    # Create HTTP response with Excel file
    response = HttpResponse(excel_file, content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    response['Content-Disposition'] = 'attachment; filename="form_data.xlsx"'
    
    return response

def export_excel(request, stp_id):
    if request.method == 'POST' and request.FILES.get('pdf'):
        pdf_file = request.FILES['pdf']
        is_pdf = pdf_file.name.split('.')[-1].lower() == 'pdf'

        if not is_pdf:
            messages.warning(request, 'Please insert PDF file only!')
            return HttpResponseRedirect(reverse('export-excel', args=[stp_id]))
        
        # Save the uploaded PDF file temporarily
        temp_dir = os.path.join(settings.BASE_DIR, 'temp')
        os.makedirs(temp_dir, exist_ok=True)
        temp_pdf_path = os.path.join(temp_dir, pdf_file.name)
        
        with open(temp_pdf_path, 'wb') as temp_pdf:
            for chunk in pdf_file.chunks():
                temp_pdf.write(chunk)
        
        # Convert PDF to Excel
        workbook, _ = pdf_to_excel(temp_pdf_path)
        
        # Remove the temporary PDF file
        os.remove(temp_pdf_path)
        
        # Create HTTP response with Excel file
        response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        response['Content-Disposition'] = f'attachment; filename="{pdf_file.name.split(".")[0]}.xlsx"'
        
        # Save the workbook to response
        excel_file = io.BytesIO()
        workbook.save(excel_file)
        excel_file.seek(0)
        response.write(excel_file.getvalue())
        
        return response
    return redirect('form', stp_id=stp_id)

def custom_404(request, exception):
    return render(request, '404.html', status=404)

def custom_500(request):
    return render(request, '500.html', status=500)

