import fitz  # PyMuPDF
import io
from openpyxl import Workbook
from openpyxl.styles import Font

def pdf_to_excel(pdf_file):
    pdf_document = fitz.open(pdf_file)
    
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "PDF Data"
    
    bold_font = Font(bold=True)
    
    worksheet.cell(row=1, column=1, value="Field").font = bold_font
    worksheet.cell(row=1, column=2, value="Value").font = bold_font
    
    current_row = 2
    
    stp_id = None  # To store the STP ID found in the document
    
    for page_num in range(len(pdf_document)):
        page = pdf_document[page_num]
        text = page.get_text("text")
        start = text.index('\n')
        text = text[start+1:]
        lines = text.split("\n")
        
        key = None
        value_lines = []
        result = []

        for line in lines:
            if line.strip() == '': 
                continue
            if ':' in line and line.strip().endswith(':'):  
                if key is not None:
                    result.append([key, ' '.join(value_lines).strip()])
                
                key = line.split(':', 1)[0].strip()
                value_lines = []
            else:
                value_lines.append(line.strip())

        if key is not None:
            result.append([key, ' '.join(value_lines).strip()])

        # Check if there's an STP ID in the result
        for item in result:
            if item[0] == 'STP ID':
                stp_id = item[1]
                break

        # Adding data to the excel sheet
        for item in result:
            worksheet.cell(row=current_row, column=1, value=item[0])
            worksheet.cell(row=current_row, column=2, value=item[1])
            current_row += 1

    return workbook, stp_id